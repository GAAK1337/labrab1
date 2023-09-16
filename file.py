import pandas as pd
import wget
import numpy
from numpy import unique
from openpyxl import load_workbook
wget.download("https://drive.google.com/uc?export=download&id=1Bo5Oili5dAvWDSzAZXjzgjS71IrmLWun")
data =pd.read_excel("lab_pi_101.xlsx")
df=pd.DataFrame(data)
kolvostr=len(df)
wb = load_workbook('lab_pi_101.xlsx')
sheet = wb.get_sheet_by_name('Лист1')
kol=0
for i in range(1, kolvostr):
    gruppa=(sheet.cell(row=i, column=11).value)
    if gruppa=='ПИ101':
        kol=kol+1
nomer=pd.unique(df[['Личный номер студента']]. values.ravel ())
nom=list()
for i in range(1, kolvostr):
    
    n=(sheet.cell(row=i, column=10).value)
    g=(sheet.cell(row=i, column=11).value)
    if g=='ПИ101':
       nom.append(n)
stud = ", ".join(str(x) for x in (unique(nom)))
uniques=pd.unique(df[['Уровень контроля']]. values.ravel ())
formi = ", ".join(uniques)
god=pd.unique(df[['Год']]. values.ravel ())
goda = ", ".join(str(x) for x in god)
print('В исходном датасете содержалось:',kolvostr,'оценок, из них:',kol, 'оценок относятся к группе ПИ101 \n В датасете находятся оценки',len (nomer), 'студентов со следующими личными номерами из группы ПИ101:', stud,'\nИспользуемые формы контроля:', formi, '\nДанные представлены по следующим учебным годам:', goda)
